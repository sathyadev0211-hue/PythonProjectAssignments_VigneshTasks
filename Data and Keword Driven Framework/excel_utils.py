# utils/excel_utils.py
# Utility module for reading test data from and writing results to Excel
# Uses openpyxl for Excel operations — part of Data Driven Testing Framework (DDTF)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


# ─── Constants ────────────────────────────────────────────────────────────────
EXCEL_FILE_PATH = "test_data.xlsx"
SHEET_NAME      = "LoginTestData"

# Column indices (1-based) matching the Excel header layout:
# Test ID | Username | Password | Date | Time of Test | Name of Tester | Test Result
COL_TEST_ID      = 1
COL_USERNAME     = 2
COL_PASSWORD     = 3
COL_DATE         = 4
COL_TIME         = 5
COL_TESTER_NAME  = 6
COL_TEST_RESULT  = 7

# Styling for result cells
PASS_FILL  = PatternFill("solid", start_color="C6EFCE")  # Light green
FAIL_FILL  = PatternFill("solid", start_color="FFC7CE")  # Light red
PASS_FONT  = Font(name="Arial", size=10, bold=True, color="276221")
FAIL_FONT  = Font(name="Arial", size=10, bold=True, color="9C0006")
CENTER     = Alignment(horizontal="center", vertical="center")


def get_test_data() -> list[dict]:
    """
    Read all test rows from the Excel file.

    Returns:
        List of dicts, each with keys:
            row_number, test_id, username, password,
            date, time, tester_name, test_result
    """
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    ws = wb[SHEET_NAME]

    test_rows = []
    # Row 1 is headers; data starts at row 2
    for row in ws.iter_rows(min_row=2, values_only=False):
        test_id     = row[COL_TEST_ID     - 1].value
        username    = row[COL_USERNAME    - 1].value
        password    = row[COL_PASSWORD    - 1].value
        date_val    = row[COL_DATE        - 1].value
        time_val    = row[COL_TIME        - 1].value
        tester_name = row[COL_TESTER_NAME - 1].value

        # Skip completely empty rows
        if not test_id:
            continue

        test_rows.append({
            "row_number":  row[0].row,     # Actual Excel row number for write-back
            "test_id":     str(test_id),
            "username":    str(username) if username else "",
            "password":    str(password) if password else "",
            "date":        str(date_val)  if date_val  else datetime.now().strftime("%d-%m-%Y"),
            "time":        str(time_val)  if time_val  else datetime.now().strftime("%I:%M %p"),
            "tester_name": str(tester_name) if tester_name else "",
            "test_result": "",
        })

    wb.close()
    return test_rows


def write_test_result(row_number: int, result: str):
    """
    Write 'Passed' or 'Failed' into the Test Result column for a given row,
    and apply colour-coded formatting.

    Args:
        row_number: The Excel row number (e.g., 2 for the first data row)
        result:     'Passed' or 'Failed'
    """
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    ws = wb[SHEET_NAME]

    result_cell = ws.cell(row=row_number, column=COL_TEST_RESULT)
    result_cell.value     = result
    result_cell.alignment = CENTER

    if result == "Passed":
        result_cell.fill = PASS_FILL
        result_cell.font = PASS_FONT
    else:
        result_cell.fill = FAIL_FILL
        result_cell.font = FAIL_FONT

    wb.save(EXCEL_FILE_PATH)
    wb.close()
